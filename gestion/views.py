from django.shortcuts import redirect, render
from .models import Usuario, Asistencia, Grupo
from .forms import PagoForm
from datetime import datetime, date
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
import csv
from django.utils.encoding import smart_str
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from .models import SessionDay
from typing import cast
import re
import json

# Create your views here.


def get_current_week():
    return datetime.now().isocalendar()

@login_required
def index(request):
    # Build display name from first and last name (fall back to full_name or username)
    fname = (request.user.first_name or "").strip()
    lname = (request.user.last_name or "").strip()
    display_name = (f"{fname} {lname}".strip())
    if not display_name:
        # try get_full_name or username
        display_name = request.user.get_full_name() or request.user.username

    # Users per role (use roles defined in the Usuario model)
    roles = [
        ("ALUMNO", "Alumnos"),
        ("ENTRENADOR", "Entrenadores"),
        ("ASISTENTE", "Asistentes"),
        ("ADMINISTRADOR", "Administradores"),
    ]
    users_labels = [label for _, label in roles]
    users_counts = [Usuario.objects.filter(rol=role).count() for role, _ in roles]

    # Sessions per day for current ISO week (Mon..Sun)
    weeknum = get_current_week()[1]
    current_year = date.today().year
    try:
        days = [date.fromisocalendar(current_year, weeknum, d) for d in range(1, 8)]
    except Exception:
        days = []
    sessions_counts = [SessionDay.objects.filter(fecha=d, active=True).count() for d in days]

    context = {
        "user_display_name": display_name,
        # JSON-encode arrays so the template can inject them directly into JS
        "users_labels_json": json.dumps(users_labels, ensure_ascii=False),
        "users_counts_json": json.dumps(users_counts),
        "sessions_labels_json": json.dumps(["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"], ensure_ascii=False),
        "sessions_counts_json": json.dumps(sessions_counts),
    }
    return render(request, "index.html", context)


@login_required
def general_table(request, grupo):
    current_week = get_current_week()[
        1
    ]  # get_current_week returns a tuple; second element is the week number
    current_year = date.today().year
    weeks = [
        {
            "week": week,
            "url": reverse("tabla_semanal", kwargs={"grupo":grupo, "semana": str(week)}),
            "start_date": date.fromisocalendar(current_year, week, 1).strftime(
                "%d/%m/%Y"
            ),
            "end_date": date.fromisocalendar(current_year, week, 7).strftime(
                "%d/%m/%Y"
            ),
        }
        for week in range(1, current_week + 1)
    ]
    return render(request, "asistencias/general_table.html", {"weeks": weeks})


@login_required
def weekly_table(request, grupo, semana):
    week = int(semana)
    current_year = date.today().year
    days = [date.fromisocalendar(current_year, week, d) for d in range(1, 8)]
    return render(
        request, "asistencias/weekly_table.html", {"week": week, "days": days, "grupo":grupo}
    )


@login_required
def view_groups(request):
    grupos = Grupo.objects.all()
    return render(request, "asistencias/grupos_table.html", {"grupos": grupos})


@login_required
def daily_attendance(request, grupo, fecha=None):
    # print(grupo)
    if not fecha:
        fecha = date.today()
    else:
        fecha = datetime.strptime(fecha, "%Y-%m-%d").date()

    estudiantes = Usuario.objects.filter(rol="ALUMNO", grupo=grupo).order_by("first_name", "last_name")
    asistencia = Asistencia.objects.filter(fecha=fecha)

    estudiantes_filtered = []
    for estudiante in estudiantes:
        # Filtrar usuarios no registrados aún o que ya pasaron su fecha de inactividad.
        if estudiante.date_joined.date() > fecha:
            continue
        if estudiante.inactivo_desde and fecha > estudiante.inactivo_desde:
            continue
        try:
            asistencia_record = asistencia.get(alumno=estudiante)
            setattr(estudiante, "asistencia", asistencia_record)
            # print("asistencia: ", asistencia_record)
        except Asistencia.DoesNotExist:
            setattr(estudiante, "asistencia", None)
        estudiantes_filtered.append(estudiante)

    context = {"estudiantes": estudiantes_filtered, "fecha": fecha}
    # Incluir 'grupo' en el contexto para permitir acciones relacionadas al grupo (por ejemplo descarga diaria)
    context["grupo"] = grupo
    # Indicar si la sesión de este grupo/fecha está activa
    try:
        session_active = SessionDay.objects.filter(grupo__pk=grupo, fecha=fecha, active=True).exists()
    except Exception:
        session_active = False
    context["session_active"] = session_active

    return render(request, "asistencias/daily_table.html", context)


@login_required
@require_http_methods(["POST"])
def htmx_create_asistencia(request, pk):
    fecha = request.POST.get("fecha")
    alumno = Usuario.objects.get(id=pk)
    # Crea asistencia con presente=True
    asistencia = Asistencia.objects.create(alumno=alumno, fecha=fecha, presente=True)
    # Render and return the updated student row
    estudiante = alumno
    setattr(estudiante, "asistencia", asistencia)
    html = render(request, "asistencias/_student_row.html", {"student": estudiante, "fecha": fecha, "session_active": True})
    return HttpResponse(html)


@login_required
@require_http_methods(["POST"])
def htmx_update_asistencia(request, pk):
    try:
        asistencia_record = Asistencia.objects.get(pk=pk)
        # Alterna el estado de 'presente'
        asistencia_record.presente = not asistencia_record.presente
        asistencia_record.save()
        estudiante = asistencia_record.alumno
        setattr(estudiante, "asistencia", asistencia_record)
        fecha = asistencia_record.fecha
        html = render(request, "asistencias/_student_row.html", {"student": estudiante, "fecha": fecha, "session_active": True})
        return HttpResponse(html)
    except Asistencia.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Asistencia not found"}, status=404)


@login_required
@require_http_methods(["POST"])
def htmx_delete_asistencia(request, pk):
    try:
        asistencia_record = Asistencia.objects.get(pk=pk)
        estudiante = asistencia_record.alumno
        fecha = asistencia_record.fecha
        asistencia_record.delete()
        # Return a row representing the student without asistencia (so the row can be replaced)
        setattr(estudiante, "asistencia", None)
        html = render(request, "asistencias/_student_row.html", {"student": estudiante, "fecha": fecha, "session_active": True})
        return HttpResponse(html)
    except Asistencia.DoesNotExist:
        return HttpResponse("Not found", status=404)



@login_required
@require_POST
def activate_session_day(request):
    """HTMX endpoint to activate the session for a given group and date.
    Expects POST params: grupo (id) and fecha (YYYY-MM-DD).
    """
    grupo_id = request.POST.get("grupo")
    fecha = request.POST.get("fecha")
    if not grupo_id or not fecha:
        return HttpResponse("Missing params", status=400)
    try:
        from datetime import datetime
        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
    except Exception:
        return HttpResponse("Fecha inválida", status=400)
    try:
        grupo = Grupo.objects.get(pk=grupo_id)
    except Grupo.DoesNotExist:
        return HttpResponse("Grupo no encontrado", status=404)

    session, created = SessionDay.objects.get_or_create(grupo=grupo, fecha=fecha_obj)
    session.active = True
    session.save()
    # Render and return the updated fragment so HTMX can swap it in-place
    # Rebuild the estudiantes/context for the fragment
    estudiantes = Usuario.objects.filter(rol="ALUMNO", grupo=grupo).order_by("first_name", "last_name")
    asistencia_qs = Asistencia.objects.filter(fecha=fecha_obj)
    estudiantes_filtered = []
    for estudiante in estudiantes:
        if estudiante.date_joined.date() > fecha_obj:
            continue
        if estudiante.inactivo_desde and fecha_obj > estudiante.inactivo_desde:
            continue
        try:
            asistencia_record = asistencia_qs.get(alumno=estudiante)
            setattr(estudiante, "asistencia", asistencia_record)
        except Asistencia.DoesNotExist:
            setattr(estudiante, "asistencia", None)
        estudiantes_filtered.append(estudiante)

    context = {
        "estudiantes": estudiantes_filtered,
        "fecha": fecha_obj,
        "grupo": grupo.pk,
        "session_active": True,
    }
    html = render(request, "asistencias/_daily_table_fragment.html", context)
    return HttpResponse(html)


@login_required
@require_POST
def deactivate_session_day(request):
    """HTMX endpoint to deactivate the session for a given group and date.
    Expects POST params: grupo (id) and fecha (YYYY-MM-DD).
    Only accessible by admin users or superusers via the template visibility.
    """
    grupo_id = request.POST.get("grupo")
    fecha = request.POST.get("fecha")
    if not grupo_id or not fecha:
        return HttpResponse("Missing params", status=400)
    try:
        from datetime import datetime
        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
    except Exception:
        return HttpResponse("Fecha inválida", status=400)
    try:
        grupo = Grupo.objects.get(pk=grupo_id)
    except Grupo.DoesNotExist:
        return HttpResponse("Grupo no encontrado", status=404)

    try:
        session = SessionDay.objects.get(grupo=grupo, fecha=fecha_obj)
        session.active = False
        session.save()
    except SessionDay.DoesNotExist:
        # nothing to do if it doesn't exist
        pass

    # Rebuild context and return fragment
    estudiantes = Usuario.objects.filter(rol="ALUMNO", grupo=grupo).order_by("first_name", "last_name")
    asistencia_qs = Asistencia.objects.filter(fecha=fecha_obj)
    estudiantes_filtered = []
    for estudiante in estudiantes:
        if estudiante.date_joined.date() > fecha_obj:
            continue
        if estudiante.inactivo_desde and fecha_obj > estudiante.inactivo_desde:
            continue
        try:
            asistencia_record = asistencia_qs.get(alumno=estudiante)
            setattr(estudiante, "asistencia", asistencia_record)
        except Asistencia.DoesNotExist:
            setattr(estudiante, "asistencia", None)
        estudiantes_filtered.append(estudiante)

    context = {
        "estudiantes": estudiantes_filtered,
        "fecha": fecha_obj,
        "grupo": grupo.pk,
        "session_active": False,
    }
    html = render(request, "asistencias/_daily_table_fragment.html", context)
    return HttpResponse(html)


@login_required
def registrar_pago(request):
    if request.method == "POST":
        form = PagoForm(request.POST)
        if form.is_valid():
            form.save()
            return render(request, "pagos/registro_exitoso.html")
    else:
        form = PagoForm()
    return render(request, "pagos/registrar_pago.html", {"form": form})


@login_required
def download_attendance_summary(request, grupo):
    """
    Genera y devuelve un CSV resumido de asistencias para todos los alumnos de un grupo.
    Columnas: nombre, fecha_primera, fecha_ultima, total_sesiones, asistencias
    """
    # Obtener estudiantes del grupo
    estudiantes = Usuario.objects.filter(rol="ALUMNO", grupo=grupo).order_by("first_name", "last_name")

    # Construir datos de resumen
    rows = []
    for estudiante in estudiantes:
        asistencias = Asistencia.objects.filter(alumno=estudiante).order_by("fecha")
        total_sesiones = asistencias.count()
        asistencias_presentes = asistencias.filter(presente=True).count()
        first = asistencias.first()
        last = asistencias.last()
        fecha_primera = first.fecha if first else ""
        fecha_ultima = last.fecha if last else ""
        rows.append({
            "nombre": estudiante.nombre_completo() if hasattr(estudiante, 'nombre_completo') else f"{estudiante.first_name} {estudiante.last_name}",
            "fecha_primera": fecha_primera,
            "fecha_ultima": fecha_ultima,
            "total_sesiones": total_sesiones,
            "asistencias": asistencias_presentes,
        })

    # Crear respuesta CSV
    response = HttpResponse(content_type="text/csv")
    filename = f"asistencias_grupo_{grupo}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        "Nombres",
        "fecha_primera",
        "fecha_ultima",
        "total_sesiones",
        "Asistencias",
    ])
    for r in rows:
        writer.writerow([
            smart_str(r["nombre"]),
            r["fecha_primera"],
            r["fecha_ultima"],
            r["total_sesiones"],
            r["asistencias"],
        ])

    return response


@login_required
def download_weekly_attendance_summary(request, grupo, semana):
    """
    Genera un CSV resumido de asistencias para un grupo en una semana ISO dada.
    Columnas: nombre, semana, fecha_inicio, fecha_fin, total_sesiones, asistencias
    """
    from datetime import date
    week = int(semana)
    current_year = date.today().year

    # calcular días de la semana como objetos date
    try:
        all_days = [date.fromisocalendar(current_year, week, d) for d in range(1, 8)]
        start_date = all_days[0]
        end_date = all_days[-1]
    except Exception:
        return HttpResponse("Semana inválida", status=400)

    # Filtrar solo los días que estén activos en SessionDay para este grupo
    active_days_qs = SessionDay.objects.filter(grupo__pk=grupo, fecha__gte=start_date, fecha__lte=end_date, active=True)
    active_days = sorted([sd.fecha for sd in active_days_qs])
    # Si no hay días activos en la semana, devolver un CSV/XLSX vacío con solo encabezado
    days = active_days

    estudiantes = Usuario.objects.filter(rol="ALUMNO", grupo=grupo).order_by("first_name", "last_name")

    # Prefetch asistencias para la semana en un solo query y construir un mapa (alumno, fecha) -> presente
    asistencias_qs = Asistencia.objects.filter(alumno__in=estudiantes, fecha__gte=start_date, fecha__lte=end_date)
    asist_map = {(a.alumno.pk, a.fecha): a.presente for a in asistencias_qs}

    # Encabezado: nombre, <dia1>, <dia2>, ..., total_sesiones, asistencias
    # Usar formato dd-mm-yyyy
    day_headers = [d.strftime("%d-%m-%Y") for d in days]
    header = ["Nombres"] + day_headers + ["total_sesiones", "Asistencias"]

    fmt = request.GET.get("format", "").lower()
    if fmt == "xlsx":
        # Generar XLSX
        wb = Workbook()
        ws = cast(Worksheet, wb.active)

        # Construir filas en memoria primero
        rows_data = []
        for estudiante in estudiantes:
            presentes_count = 0
            total_records = 0
            name = estudiante.nombre_completo() if hasattr(estudiante, 'nombre_completo') else f"{estudiante.first_name} {estudiante.last_name}"
            day_values = []
            for d in days:
                presente = asist_map.get((estudiante.pk, d), False)
                if (estudiante.pk, d) in asist_map:
                    total_records += 1
                if presente:
                    presentes_count += 1
                    day_values.append("Presente")
                else:
                    day_values.append("Ausente")
            row_values = [smart_str(name)] + day_values + [total_records, presentes_count]
            rows_data.append(row_values)

        # Escribir encabezado y filas (solo una vez)
        ws.append(header)
        for r in rows_data:
            ws.append(r)

        # Ajustar ancho de columnas según el contenido
        nrows = len(rows_data) + 1
        ncols = len(header)
        for col_idx in range(1, ncols + 1):
            # Calcular ancho en base a las filas (rows_data) y asegurarse de que el header también se considera
            header_len = len(str(header[col_idx - 1]))
            max_row_len = 0
            for r in rows_data:
                try:
                    val = r[col_idx - 1]
                except Exception:
                    val = ""
                l = len(str(val))
                if l > max_row_len:
                    max_row_len = l
            max_length = max(header_len, max_row_len)
            # Añadir padding extra para que no quede demasiado pegado
            adjusted_width = max_length + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Resaltar celdas con 'Ausente' en rojo
        red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=nrows, min_col=2, max_col=ncols):
            for cell in row:
                try:
                    if str(cell.value).strip().lower() == "ausente":
                        cell.fill = red_fill
                except Exception:
                    pass

        # Crear tabla Excel con estilo
        last_col = get_column_letter(ncols)
        table_range = f"A1:{last_col}{nrows}"
        # Sanitize table name: only word chars and underscores
        table_name = f"T_AsistenciasSemana_{week}_{grupo}"
        table_name = re.sub(r"\W+", "_", table_name)
        tab = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        xlsx_filename = f"asistencias_grupo_{grupo}_semana_{week}.xlsx"
        xlsx_response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        xlsx_response["Content-Disposition"] = f'attachment; filename="{xlsx_filename}"'
        return xlsx_response
    else:
        # Generar CSV por defecto
        response = HttpResponse(content_type="text/csv")
        filename = f"asistencias_grupo_{grupo}_semana_{week}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(header)

        for estudiante in estudiantes:
            presentes_count = 0
            total_records = 0
            name = estudiante.nombre_completo() if hasattr(estudiante, 'nombre_completo') else f"{estudiante.first_name} {estudiante.last_name}"
            day_values = []
            for d in days:
                presente = asist_map.get((estudiante.pk, d), False)
                if (estudiante.pk, d) in asist_map:
                    total_records += 1
                if presente:
                    presentes_count += 1
                    day_values.append("Presente")
                else:
                    day_values.append("Ausente")
            writer.writerow([smart_str(name)] + day_values + [total_records, presentes_count])

        return response


@login_required
def download_daily_attendance_summary(request, grupo, fecha):
    """
    Genera un CSV con el resumen de asistencias para una fecha específica y grupo.
    Columnas: nombre, fecha, presente, nota
    """
    from datetime import datetime
    try:
        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
    except Exception:
        return HttpResponse("Fecha inválida", status=400)

    # Verificar si la sesión para este grupo y fecha está activa
    try:
        session_active = SessionDay.objects.filter(grupo__pk=grupo, fecha=fecha_obj, active=True).exists()
    except Exception:
        session_active = False
    if not session_active:
        return HttpResponse(
            "La sesión para esta fecha no está activa. Active la sesión para incluirla en los reportes.",
            status=404,
        )

    estudiantes = Usuario.objects.filter(rol="ALUMNO", grupo=grupo).order_by("first_name", "last_name")
    rows = []
    for estudiante in estudiantes:
        try:
            asistencia = Asistencia.objects.get(alumno=estudiante, fecha=fecha_obj)
            presente = asistencia.presente
            nota = asistencia.nota
        except Asistencia.DoesNotExist:
            presente = False
            nota = ""
        rows.append(
            {
                "nombre": estudiante.nombre_completo()
                if hasattr(estudiante, "nombre_completo")
                else f"{estudiante.first_name} {estudiante.last_name}",
                "fecha": fecha_obj,
                "presente": presente,
            }
        )

    fmt = request.GET.get("format", "").lower()
    if fmt == "xlsx":
        # Generar XLSX con encabezado como tabla y ajuste de anchos
        wb = Workbook()
        ws = cast(Worksheet, wb.active)

        header = ["Nombres", "Fecha", "Asistencias"]

        # Construir filas en memoria
        rows_data = []
        for r in rows:
            presente_str = "Presente" if r["presente"] else "Ausente"
            fecha_str = (
                r["fecha"].strftime("%d-%m-%Y")
                if hasattr(r["fecha"], "strftime")
                else str(r["fecha"])
            )
            rows_data.append([smart_str(r["nombre"]), fecha_str, presente_str])

        # Escribir encabezado y filas (solo una vez)
        ws.append(header)
        for row_item in rows_data:
            ws.append(row_item)

        # Ajustar ancho de columnas según el contenido
        nrows = len(rows_data) + 1
        ncols = len(header)
        for col_idx in range(1, ncols + 1):
            header_len = len(str(header[col_idx - 1]))
            max_row_len = 0
            for row_item in rows_data:
                try:
                    val = row_item[col_idx - 1]
                except Exception:
                    val = ""
                l = len(str(val))
                if l > max_row_len:
                    max_row_len = l
            max_length = max(header_len, max_row_len)
            adjusted_width = max_length + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Resaltar celdas con 'Ausente' en rojo
        red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
        for row_cells in ws.iter_rows(min_row=2, max_row=nrows, min_col=2, max_col=1 + ncols):
            for cell in row_cells:
                try:
                    if str(cell.value).strip().lower() == "ausente":
                        cell.fill = red_fill
                except Exception:
                    pass

        # Crear tabla Excel con estilo
        last_col = get_column_letter(ncols)
        table_range = f"A1:{last_col}{nrows}"
        # Sanitize table name and include group + date (YYYYMMDD) to keep it unique and valid
        try:
            date_part = fecha_obj.strftime("%Y%m%d")
        except Exception:
            date_part = str(fecha).replace("-", "_")
        table_name = f"T_AsistenciasDiaria_{grupo}_{date_part}"
        table_name = re.sub(r"\W+", "_", table_name)
        tab = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        xlsx_filename = f"asistencias_grupo_{grupo}_{fecha}.xlsx"
        xlsx_response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        xlsx_response["Content-Disposition"] = f'attachment; filename="{xlsx_filename}"'
        return xlsx_response
    else:
        # Generar CSV por defecto con 'Presente'/'Ausente'
        response = HttpResponse(content_type="text/csv")
        filename = f"asistencias_grupo_{grupo}_{fecha}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(["Nombres", "Fecha", "Asistencias"])
        for r in rows:
            presente_str = "Presente" if r["presente"] else "Ausente"
            fecha_str = (
                r["fecha"].strftime("%d-%m-%Y")
                if hasattr(r["fecha"], "strftime")
                else str(r["fecha"])
            )
            writer.writerow([smart_str(r["nombre"]), fecha_str, presente_str])
        return response
